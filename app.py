from flask import Flask, render_template, request, send_file, redirect, flash, session,url_for
import pandas as pd
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
import pytz
from flask_mysqldb import MySQL,MySQLdb # pip install Flask-MySQLdb
import os
from werkzeug.utils import secure_filename
import jinja2
import hashlib # ENMASCARA URLS


app = Flask(__name__, template_folder='template')
app.secret_key = "pinchellave"
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_PORT'] = 3306  # Puerto específico
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''
app.config['MYSQL_DB'] = 'proyecto'

app.config['UPLOAD_FOLDER'] = os.path.abspath(os.path.join(os.path.dirname(__file__), 'archivos'))
app.config['MYSQL_CURSORCLASS'] = 'DictCursor'
mysql = MySQL(app)


print("Intentando conectar a la base de datos...")
@app.route('/verificar-conexion')
def verificar_conexion():
    try:
        # Intenta ejecutar una consulta simple
        with mysql.connection.cursor() as cursor:
            cursor.execute("SELECT 1")
            result = cursor.fetchone()

        if result:
            return "Conexión exitosa a la base de datos."
        print("Conexión exitosa a la base de datos.")

    except Exception as e:
        return f"Error al intentar conectarse a la base de datos: {str(e)}"

def procesar_archivo(archivo,titulo):
    datos = pd.read_excel(archivo)

    # ... (Resto de tu código de procesamiento)
    # Almacenar los últimos valores de las columnas 'TOTAL' y 'Unnamed: 27'
    ultimo_valor_total = datos['TOTAL'].iloc[-1]
    ultimo_valor_unnamed = datos['Unnamed: 27'].iloc[-1]

    # Seleccionar solo la columna "Descripción"
    columnas_seleccionadas = datos[['Descripción', 'TOTAL', 'Unnamed: 27']]
    columnas_seleccionadas = columnas_seleccionadas.reset_index(drop=True)

    # Convertir la columna 'Unnamed: 27' a tipo numérico solo en las filas numéricas
    columnas_seleccionadas.iloc[1:, 2] = pd.to_numeric(columnas_seleccionadas.iloc[1:, 2], errors='coerce')

    # Filtrar filas que contienen 'total' en la columna 'Descripción', excluyendo la fila 0
    columnas_filtradas = columnas_seleccionadas.copy()
    columnas_filtradas.iloc[1:, 2] = pd.to_numeric(columnas_filtradas.iloc[1:, 2], errors='coerce')
    columnas_filtradas = columnas_filtradas[~columnas_filtradas['Descripción'].fillna('').str.lower().str.contains('total')]
    
    # Ordenar solo las filas que contienen datos numéricos por la columna 'Unnamed: 27' de mayor a menor
    columnas_ordenadas = columnas_filtradas.iloc[1:].sort_values(by='Unnamed: 27', ascending=False)

    # Seleccionar los primeros 10 datos después de ordenar
    primeros_10_datos = columnas_ordenadas.head(10)

    
    # Concatenar la fila 0 y los primeros 10 datos ordenados
    resultado_final = pd.concat([columnas_filtradas.iloc[:1], primeros_10_datos])

    # Calcular la suma total de 'TOTAL' y 'Unnamed: 27' excluyendo la fila 0
    suma_total_TOTAL = resultado_final.iloc[1:]['TOTAL'].sum()
    suma_total_Unnamed = resultado_final.iloc[1:]['Unnamed: 27'].sum()

    # Crear una nueva fila con el valor "Otros" y las sumas calculadas
    fila_otros = pd.DataFrame({"Descripción": ["Otros"], "TOTAL": [suma_total_TOTAL], "Unnamed: 27": [suma_total_Unnamed]})

    # Concatenar la fila "Otros" al final del DataFrame
    resultado_final = pd.concat([resultado_final, fila_otros], ignore_index=True)
    
    fila_total = pd.DataFrame({
    "Descripción": ["Total"],
    "TOTAL": [ultimo_valor_total],
    "Unnamed: 27": [ultimo_valor_unnamed]
    })

    # Concatenar la fila "Total" al final del DataFrame
    resultado_final = pd.concat([resultado_final, fila_total], ignore_index=True)

# Calcular las diferencias entre las filas 'Total' y 'Otros'
    diferencia_total_otros_total = resultado_final.loc[resultado_final['Descripción'] == 'Total', 'TOTAL'].values[0] - resultado_final.loc[resultado_final['Descripción'] == 'Otros', 'TOTAL'].values[0]
    diferencia_total_otros_unnamed = resultado_final.loc[resultado_final['Descripción'] == 'Total', 'Unnamed: 27'].values[0] - resultado_final.loc[resultado_final['Descripción'] == 'Otros', 'Unnamed: 27'].values[0]
    
    # Actualizar los valores en la fila 'Otros'
    resultado_final.loc[resultado_final['Descripción'] == 'Otros', 'TOTAL'] = diferencia_total_otros_total
    resultado_final.loc[resultado_final['Descripción'] == 'Otros', 'Unnamed: 27'] = diferencia_total_otros_unnamed
    # ... (Código para generar el resultado_final)

   

    
# Cambiar el nombre de la columna "Unnamed: 27" a "TOTAL"
    resultado_final = resultado_final.rename(columns={'Unnamed: 27': 'TOTAL'})

    # Crear un nuevo libro de trabajo de Excel
    workbook = Workbook()
    sheet = workbook.active

    # Agregar un título encima de la tabla con la fecha actual
    tz = pytz.timezone('America/La_Paz')
    fecha_actual = datetime.now(tz).strftime("%d-%m-%Y, Horas: %H:%M:%S")
    titulo_completo = f"TOP 10 DE LOS TRAMITES MAS SOLICITADOS EN EL {titulo}\n(FECHA: {fecha_actual})"
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(resultado_final.columns))
    sheet['A1'] = titulo_completo
    sheet['A1'].font = Font(size=13, bold=True)
    sheet['A1'].alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    # Ajustar el alto de la fila para adaptarse al texto del título
    sheet.row_dimensions[1].height = 57  # Ajusta el valor según tus necesidades

    # Obtener el número de filas ocupadas por los encabezados y datos
    num_filas_datos = len(resultado_final) + 2  # Se suma 2 para incluir la fila de encabezados y la fila del título

    # Copiar los encabezados del DataFrame al libro de trabajo
    headers = [col for col in resultado_final.columns]
    sheet.append(headers)

    # Copiar los datos del DataFrame al libro de trabajo
    for row in resultado_final.itertuples(index=False):
        sheet.append(list(row))

    # Aplicar bordes a todas las celdas
    for row in sheet.iter_rows(min_row=2, max_row=num_filas_datos, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = Border(left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='thin', color='000000'))

    # Aplicar color a la fila de encabezados (columnas)
    for cell in sheet[2]:
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Aplicar color a la fila de encabezados (columnas)    
    for cell in sheet[3]:
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Aplicar color a la penúltima fila de la tabla (restamos 1 al índice máximo de filas)
    for cell in sheet[num_filas_datos - 1]:
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Amarillo

    # Aplicar color a la última fila de la tabla
    for cell in sheet[num_filas_datos]:
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Ajustar el formato del texto para mostrar múltiples líneas
    for row in sheet.iter_rows(min_row=2, max_row=num_filas_datos, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    # Establecer el ancho de la columna "Descripcion"
    sheet.column_dimensions['A'].width = 40  # Ajusta el valor según tus necesidades
    sheet.column_dimensions['B'].width = 13
    sheet.column_dimensions['C'].width = 13

    # Guardar el archivo Excel
    workbook.save('reporte_top_10_go.xlsx')


    # ... (Código para dar formato y guardar el archivo)

    # Convertir el libro de trabajo en un objeto BytesIO para descargar
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output

@app.route('/')
def home():
    return render_template('index.html')


@app.route('/subir')
def subir():
    return render_template('subir.html')

@app.route('/procesar', methods=['POST'])
def procesar():
    if 'archivo' not in request.files:
        flash('No se proporcionó ningún archivo')
        return redirect(request.url)

    archivo = request.files['archivo']
    titulo = request.form.get('titulo')
    
    if archivo.filename == '':
        flash('Ningún archivo seleccionado')
        return redirect(request.url)

    if archivo:
        output = procesar_archivo(archivo,titulo)
        return send_file(output, as_attachment=True, download_name=f'TOP_10_{titulo}.xlsx')

    return render_template('index.html')

@app.route('/top_10')
def top_10():
    if 'logueado' in session:
        return render_template('top_10.html')
    else:
        return redirect(url_for('home'))

@app.route('/gestiones')
def gestiones():
    if 'logueado' in session:
        return render_template('gestiones.html')
    else:
        return redirect(url_for('home'))

#REGISTRAR TRAMITES CONSULARES
def guardar_en_base_de_datos(pais, oficina, detalle, top, gestiones, general,gestion):
    cur = mysql.connection.cursor()
    cur.execute("INSERT INTO tramites_consulares (pais, oficina_consular, detalle, top_10, gestiones, general,tram_gestion) VALUES (%s, %s, %s, %s, %s, %s,%s)",
                (pais, oficina, detalle, top, gestiones, general,gestion))
    mysql.connection.commit()
    cur.close()
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx'}

@app.route('/registro-crear', methods = ["GET", "POST"])
def registro():
    try:
        pais = request.form['txtPais']
        oficina = request.form['txtOficina']
        detalle = save_file(request.files.get('txtDetalle'))
        top = save_file(request.files.get('txtTop_10'))
        gestiones = save_file(request.files.get('txtGestion'))
        general = save_file(request.files.get('txtGeneral'))
        gestion = int(request.form['txtGestiones'])

        # Realizar una consulta para verificar si ya existe un registro con los mismos valores
        cur = mysql.connection.cursor()
        cur.execute("SELECT * FROM tramites_consulares WHERE pais = %s AND oficina_consular = %s AND tram_gestion = %s", (pais, oficina, gestion))
        existing_record = cur.fetchone()

        # Si no hay un registro existente, realizar la inserción
        if not existing_record:
            guardar_en_base_de_datos(pais, oficina, detalle, top, gestiones, general,gestion)
            return render_template('registro_tramites.html', mensaje2="Registro Exitoso")
        else:
            return render_template('registro_tramites.html', mensaje_duplicado="Ya Existe un Registro con los Mismos Datos")

    except Exception as e:
        print("Error:", str(e))
        return render_template('registro_tramites.html', mensaje_error="Error al registrar: {}".format(str(e)))


def save_file(file):
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        return filepath
    else:
        return None
    
#----VALIDACION LOGIN
@app.route('/acceso-login', methods= ["GET", "POST"])
def login():
    if request.method == 'POST' and 'txtUsuario' in request.form and 'txtPassword' in request.form:
       
        _usuario = request.form['txtUsuario']
        _password = request.form['txtPassword']

        cur = mysql.connection.cursor()
        cur.execute('SELECT * FROM usuarios WHERE Usuario = %s AND Password = %s', (_usuario, _password,))
        account = cur.fetchone()
      
        if account:
            session['logueado'] = True
            session['ID'] = account['ID']
            session['Nombre'] = account['Nombre']
            session['Usuario'] = account['Usuario']
            session['Password'] = account['Password']
            session['id_rol'] = account['id_rol']
            mensaje5="bienvenido!"

            if session['id_rol'] == 1:
                    return render_template("vista_consular.html",mensajeinicio=mensaje5)
            elif session['id_rol'] == 2:
                    return render_template("vista_consular_usuario.html",mensajeinicio=mensaje5)
            print(account[0])
        else:
         
            return render_template('index.html',mensaje="Usuario O Contraseña Incorrectas")
  
    return render_template('index.html')

#----------------------------------------------------------------- 
@app.route('/registro_tramites')
def registro_tramites():
    if 'logueado' in session:
        return render_template('registro_tramites.html')
    else:
        return redirect(url_for('home'))

#VISTA---CONSULAR-----------BUSQUEDA DE TRAMITES
@app.route('/vista_consular', methods=['POST'])
def vistas_consular():
        # Conectar a la base de datos
        cur = mysql.connection.cursor()

        # Obtener los valores del formulario
        pais = request.form['txtPais']
        oficina = request.form['txtOficina']
        gestion = int(request.form['txtGestiones'])

          # Construir la consulta SQL con JOIN entre tramites_consulares y gestiones
        sql = "SELECT t.*, g.gestion FROM tramites_consulares t JOIN gestiones g ON t.tram_gestion = g.id_gestion WHERE t.pais = %s AND t.oficina_consular = %s AND t.tram_gestion = %s"
        params = (pais, oficina, gestion)

        cur.execute(sql, params)
        result = cur.fetchall()
    # Mostrar los resultados
        cur.close()
 
        return render_template("vista_consular.html",  result=result)


@app.route('/descargar_archivo/<archivo>')
def descargar_archivo(archivo):
    directorio_archivos = app.config['UPLOAD_FOLDER']
    ruta_completa = os.path.join(directorio_archivos, archivo)

    if not os.path.isfile(ruta_completa):
        return "Archivo no encontrado", 404

    return send_file(ruta_completa, as_attachment=True)


@app.route('/vista_consular')
def vista_consular():
    if 'logueado' in session:
        return render_template('vista_consular.html')
    else:
        return redirect(url_for('home'))

 # vista CONSULAR USUARIO 2
@app.route('/top_10_usuario')
def top_10_usuario():
    if 'logueado' in session:
        return render_template('top_10_usuario.html')
    else:
        return redirect(url_for('home'))

@app.route('/vista_consular_usuario', methods=['POST'])
def vistas_consular_user():
        # Conectar a la base de datos
        cur = mysql.connection.cursor()

        # Obtener los valores del formulario
        pais = request.form['txtPais']
        oficina = request.form['txtOficina']
        gestion = int(request.form['txtGestiones'])

          # Construir la consulta SQL con JOIN entre tramites_consulares y gestiones
        sql = "SELECT t.*, g.gestion FROM tramites_consulares t JOIN gestiones g ON t.tram_gestion = g.id_gestion WHERE t.pais = %s AND t.oficina_consular = %s AND t.tram_gestion = %s"
        params = (pais, oficina, gestion)

        cur.execute(sql, params)
        result = cur.fetchall()
    # Mostrar los resultados
        cur.close()
 
        return render_template("vista_consular_usuario.html",  result=result)
    
@app.route('/vista_consular_usuario')
def vista_consular_usuario():
    if 'logueado' in session:
        # Genera un token cifrado con información del usuario
        # Genera una ruta enmascarada con la información del usuario
        ruta_enmascarada = enmascarar_ruta(f"user_{session['ID']}")
        return render_template('vista_consular_usuario.html',ruta_enmascarada=ruta_enmascarada)
    else:
        return redirect(url_for('home'))
#----------------------------------------------------

@app.route('/editar_campo/<int:id>/<campo>', methods=['GET'])
def editar_campo(id, campo):
    # Aquí debes cargar los datos del campo específico y renderizar el formulario de edición
    return render_template('formulario_edicion.html', id=id, campo=campo)

# La ruta para procesar el formulario de edición
# La ruta para procesar el formulario de edición
@app.route('/guardar_edicion/<int:id>', methods=['POST'])
def guardar_edicion(id):
    campo = request.form['campo']
    nuevo_valor = save_file(request.files.get('nuevo_valor'))
    

    # Llama a la función de guardado con el parámetro de id para la edición
    guardar_en_base_de_datoss(id, campo, nuevo_valor)
    # Redirige a la página de muestra después de editar
    return render_template('vista_consular.html',updates="Registro Actualizado")


# La función para guardar en la base de datos
def guardar_en_base_de_datoss(id, campo, nuevo_valor):
    # Conectar a la base de datos
    cur = mysql.connection.cursor()

    # Construir la consulta SQL para actualizar el campo específico
    sql = ""
    if campo == 'detalle':
        sql = "UPDATE tramites_consulares SET detalle = %s WHERE id = %s"
    elif campo == 'top_10':
        sql = "UPDATE tramites_consulares SET top_10 = %s WHERE id = %s"
    elif campo == 'general':
        sql = "UPDATE tramites_consulares SET general = %s WHERE id = %s"
    elif campo == 'gestion':
        sql = "UPDATE tramites_consulares SET gestion = %s WHERE id = %s"

    try:
        # Imprime la consulta SQL para depuración
        print("SQL:", cur.mogrify(sql, (nuevo_valor, id)))

        # Ejecutar la consulta
        if nuevo_valor is not None and id is not None:
            cur.execute(sql, (nuevo_valor, id))
        else:
            print("Alguno de los valores es nulo")

        # Confirmar la transacción
        mysql.connection.commit()

    except Exception as e:
        # Manejar el error e imprimirlo
        print("Error al ejecutar la consulta:", str(e))

    finally:
        # Cerrar el cursor
        cur.close()

# Ruta protegida que requiere autenticación
@app.route('/pagina_protegida')
def pagina_protegida():
    if 'logueado' in session:
        return f'Hola, {session["Nombre"]}!'
    else:
        return redirect(url_for('login'))
#destruir sesion
@app.route('/cerrar-sesion')
def logout():
    session.clear()
    return redirect(url_for('home'))

# Función para generar una ruta enmascarada usando MD5
def enmascarar_ruta(original):
    return hashlib.md5(original.encode()).hexdigest()


#------------------------------------------------------------------------------
if __name__ == '__main__':
    app.secret_key = "pinchellave"
    app.run(debug=True, host='0.0.0.0', port=5000, threaded=True)